import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from db_config import (
    get_postgres_connection,
    print_header,
    get_output_path
)

print_header("PostgreSQL Table Details Viewer")

# Connect to PostgreSQL
print("\nConnecting to PostgreSQL...")
pg_conn = get_postgres_connection()
print("✓ PostgreSQL Connected")

# Get table name from user
print("\n" + "=" * 80)
table_name = input("Enter PostgreSQL table name (e.g., event_master): ")
print("=" * 80)

try:
    print(f"\nFetching details for table: {table_name}")
    
    cursor = pg_conn.cursor()
    
    # Get column details
    query = f"""
    SELECT 
        column_name,
        data_type,
        character_maximum_length,
        CASE 
            WHEN is_nullable = 'YES' THEN 'NULL'
            ELSE 'NOT NULL'
        END as nullable,
        column_default,
        ordinal_position
    FROM information_schema.columns
    WHERE table_name = '{table_name}' 
    AND table_schema = 'public'
    ORDER BY ordinal_position
    """
    
    df = pd.read_sql(query, pg_conn)
    
    if len(df) == 0:
        print(f"\n✗ Table '{table_name}' not found in PostgreSQL!")
        
        # Show available tables
        print("\nAvailable tables:")
        cursor.execute("""
            SELECT table_name 
            FROM information_schema.tables 
            WHERE table_schema = 'public' 
            AND table_type = 'BASE TABLE'
            ORDER BY table_name
        """)
        tables = cursor.fetchall()
        for idx, table in enumerate(tables[:20], 1):
            print(f"  {idx}. {table[0]}")
        
        if len(tables) > 20:
            print(f"  ... and {len(tables) - 20} more tables")
        
        pg_conn.close()
        input("\nPress Enter to exit...")
        exit()
    
    # Get row count
    cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
    row_count = cursor.fetchone()[0]
    
    # Print results to console
    print("\n" + "=" * 80)
    print(f"Table: {table_name}")
    print("=" * 80)
    print(f"\nTotal Columns: {len(df)}")
    print(f"Total Rows: {row_count:,}")
    print("\n" + "-" * 80)
    print(f"{'#':<4} {'Column Name':<30} {'Data Type':<25} {'Nullable':<12}")
    print("-" * 80)
    
    for idx, row in df.iterrows():
        col_name = row['column_name']
        data_type = row['data_type']
        
        # Add length for character types
        if row['character_maximum_length'] and pd.notna(row['character_maximum_length']):
            data_type = f"{data_type}({int(row['character_maximum_length'])})"
        
        nullable = row['nullable']
        
        print(f"{idx+1:<4} {col_name:<30} {data_type:<25} {nullable:<12}")
    
    # Create formatted Excel output
    print("\n" + "=" * 80)
    print("Creating detailed Excel report...")
    print("=" * 80)
    
    # Prepare data for Excel
    excel_data = []
    for idx, row in df.iterrows():
        data_type = row['data_type']
        if row['character_maximum_length'] and pd.notna(row['character_maximum_length']):
            data_type = f"{data_type}({int(row['character_maximum_length'])})"
        
        excel_data.append({
            'Position': int(row['ordinal_position']),
            'Column Name': row['column_name'],
            'Data Type': data_type,
            'Nullable': row['nullable'],
            'Default Value': row['column_default'] if pd.notna(row['column_default']) else ''
        })
    
    df_excel = pd.DataFrame(excel_data)
    
    # Save to Excel with formatting
    output_file = get_output_path(f"PG_{table_name}_details_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write table details
        df_excel.to_excel(writer, sheet_name='Table_Details', index=False, startrow=3)
        
        workbook = writer.book
        worksheet = writer.sheets['Table_Details']
        
        # Add table info at top
        worksheet['A1'] = f'PostgreSQL Table: {table_name}'
        worksheet['A2'] = f'Total Columns: {len(df)}'
        worksheet['C2'] = f'Total Rows: {row_count:,}'
        
        # Style header
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        worksheet['A1'].font = header_font
        worksheet['A1'].fill = header_fill
        
        # Merge title cell
        worksheet.merge_cells('A1:E1')
        worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Style info row
        info_font = Font(bold=True, size=11)
        for cell in ['A2', 'C2']:
            worksheet[cell].font = info_font
        
        # Style column headers
        col_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        col_header_font = Font(bold=True, size=11)
        center_align = Alignment(horizontal='center', vertical='center')
        
        for cell in ['A4', 'B4', 'C4', 'D4', 'E4']:
            worksheet[cell].fill = col_header_fill
            worksheet[cell].font = col_header_font
            worksheet[cell].alignment = center_align
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 10
        worksheet.column_dimensions['B'].width = 35
        worksheet.column_dimensions['C'].width = 30
        worksheet.column_dimensions['D'].width = 15
        worksheet.column_dimensions['E'].width = 30
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=4, max_row=len(df)+4, min_col=1, max_col=5):
            for cell in row:
                cell.border = thin_border
                if cell.column == 1 or cell.column == 4:  # Position and Nullable columns
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(vertical='center')
        
        # Add sample data sheet
        print("Fetching sample data (first 10 rows)...")
        cursor.execute(f"SELECT * FROM {table_name} LIMIT 10")
        sample_data = cursor.fetchall()
        
        if sample_data:
            column_names = [desc[0] for desc in cursor.description]
            df_sample = pd.DataFrame(sample_data, columns=column_names)
            df_sample.to_excel(writer, sheet_name='Sample_Data', index=False)
            
            # Format sample data sheet
            sample_sheet = writer.sheets['Sample_Data']
            for cell in sample_sheet[1]:
                cell.fill = col_header_fill
                cell.font = col_header_font
                cell.alignment = center_align
    
    print(f"\n✓ Excel report saved to: {output_file}")
    print(f"\nSheets created:")
    print(f"  1. Table_Details - Column information")
    print(f"  2. Sample_Data - First 10 rows of data")

except Exception as e:
    print(f"\n✗ Error fetching table details!")
    print(f"Error: {str(e)}")
    import traceback
    traceback.print_exc()

finally:
    pg_conn.close()
    print("\n" + "=" * 80)
    print("Connection closed")
    print("=" * 80)

input("\nPress Enter to exit...")