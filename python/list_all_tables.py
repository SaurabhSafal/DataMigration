import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from db_config import (
    get_mssql_connection, 
    get_postgres_connection,
    print_header,
    get_output_path
)

print_header("All Tables List - MSSQL vs PostgreSQL")

# Connect to databases
print("\n[1/3] Connecting to MSSQL...")
mssql_conn = get_mssql_connection()
print("✓ MSSQL Connected")

print("\n[2/3] Connecting to PostgreSQL...")
pg_conn = get_postgres_connection()
print("✓ PostgreSQL Connected")

try:
    # Get all MSSQL tables
    print("\n[3/3] Fetching all tables...")
    
    mssql_query = """
    SELECT 
        TABLE_NAME,
        (SELECT COUNT(*) 
         FROM INFORMATION_SCHEMA.COLUMNS 
         WHERE TABLE_NAME = t.TABLE_NAME 
         AND TABLE_SCHEMA = 'dbo') as COLUMN_COUNT
    FROM INFORMATION_SCHEMA.TABLES t
    WHERE TABLE_TYPE = 'BASE TABLE' 
    AND TABLE_SCHEMA = 'dbo'
    ORDER BY TABLE_NAME
    """
    
    df_mssql = pd.read_sql(mssql_query, mssql_conn)
    print(f"✓ MSSQL: Found {len(df_mssql)} tables")
    
    # Get all PostgreSQL tables
    pg_query = """
    SELECT 
        table_name,
        (SELECT COUNT(*) 
         FROM information_schema.columns 
         WHERE table_name = t.table_name 
         AND table_schema = 'public') as column_count
    FROM information_schema.tables t
    WHERE table_type = 'BASE TABLE' 
    AND table_schema = 'public'
    ORDER BY table_name
    """
    
    df_pg = pd.read_sql(pg_query, pg_conn)
    print(f"✓ PostgreSQL: Found {len(df_pg)} tables")
    
    # Create side-by-side comparison
    max_rows = max(len(df_mssql), len(df_pg))
    
    comparison_data = []
    
    for i in range(max_rows):
        row_data = {}
        
        # MSSQL tables
        if i < len(df_mssql):
            row_data['MSSQL_TABLE_NAME'] = df_mssql.iloc[i]['TABLE_NAME']
            row_data['MSSQL_COLUMNS'] = int(df_mssql.iloc[i]['COLUMN_COUNT'])
        else:
            row_data['MSSQL_TABLE_NAME'] = ''
            row_data['MSSQL_COLUMNS'] = ''
        
        # Empty separator
        row_data['SEPARATOR'] = ''
        
        # PostgreSQL tables
        if i < len(df_pg):
            row_data['PG_table_name'] = df_pg.iloc[i]['table_name']
            row_data['PG_columns'] = int(df_pg.iloc[i]['column_count'])
        else:
            row_data['PG_table_name'] = ''
            row_data['PG_columns'] = ''
        
        comparison_data.append(row_data)
    
    df_comparison = pd.DataFrame(comparison_data)
    
    # Save to Excel with formatting
    output_file = get_output_path(f"All_Tables_Comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_comparison.to_excel(writer, sheet_name='Tables_Comparison', index=False, startrow=1)
        
        workbook = writer.book
        worksheet = writer.sheets['Tables_Comparison']
        
        # Add database names in first row
        worksheet['A1'] = 'mssql - WCL_MVC'
        worksheet['D1'] = 'postgres - navikaran_mig'
        
        # Merge cells for headers
        worksheet.merge_cells('A1:B1')
        worksheet.merge_cells('D1:E1')
        
        # Style the header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        center_align = Alignment(horizontal='center', vertical='center')
        
        for cell in ['A1', 'D1']:
            worksheet[cell].fill = header_fill
            worksheet[cell].font = header_font
            worksheet[cell].alignment = center_align
        
        # Style column headers
        col_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        col_header_font = Font(bold=True, size=11)
        
        # Set column headers
        worksheet['A2'] = 'TABLE_NAME'
        worksheet['B2'] = 'COLUMNS'
        worksheet['C2'] = ''
        worksheet['D2'] = 'table_name'
        worksheet['E2'] = 'columns'
        
        for cell in ['A2', 'B2', 'D2', 'E2']:
            worksheet[cell].fill = col_header_fill
            worksheet[cell].font = col_header_font
            worksheet[cell].alignment = center_align
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 35
        worksheet.column_dimensions['B'].width = 12
        worksheet.column_dimensions['C'].width = 3
        worksheet.column_dimensions['D'].width = 35
        worksheet.column_dimensions['E'].width = 12
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=max_rows+2, min_col=1, max_col=5):
            for cell in row:
                if cell.column != 3:  # Skip separator column
                    cell.border = thin_border
                    if cell.row > 2 and cell.column in [2, 5]:  # Number columns
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(vertical='center')
        
        # Make separator column gray
        gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for row in range(1, max_rows + 3):
            worksheet[f'C{row}'].fill = gray_fill
    
    # Print summary
    print("\n" + "=" * 80)
    print("✓ ALL TABLES COMPARISON COMPLETE!")
    print("=" * 80)
    print(f"\nSummary:")
    print(f"  MSSQL Tables: {len(df_mssql)}")
    print(f"  PostgreSQL Tables: {len(df_pg)}")
    print(f"\n✓ File saved to: {output_file}")
    
    # Show some sample tables
    print(f"\nSample MSSQL Tables (first 5):")
    for i in range(min(5, len(df_mssql))):
        print(f"  - {df_mssql.iloc[i]['TABLE_NAME']} ({df_mssql.iloc[i]['COLUMN_COUNT']} columns)")
    
    print(f"\nSample PostgreSQL Tables (first 5):")
    for i in range(min(5, len(df_pg))):
        print(f"  - {df_pg.iloc[i]['table_name']} ({df_pg.iloc[i]['column_count']} columns)")

except Exception as e:
    print(f"\n✗ Failed to fetch tables!")
    print(f"Error: {str(e)}")
    import traceback
    traceback.print_exc()

finally:
    mssql_conn.close()
    pg_conn.close()
    print("\n" + "=" * 80)
    print("Connections closed")
    print("=" * 80)

input("\nPress Enter to exit...")