#!/usr/bin/env python3
"""
compare_tables_powerful_auto_mapping_fixed.py

Fix for AttributeError: 'MergedCell' object has no attribute 'column_letter'
that occurred when writing/styling the Excel workbook.

What I changed:
- Replaced uses of `cell.column_letter` (which raises on MergedCell objects)
  with a small, safe helper that extracts the column letters from `cell.coordinate`.
  This works for normal Cell and MergedCell objects.
- Applied the helper everywhere the code previously referenced `cell.column_letter`.
- Kept all other logic and scoring identical to the previous powerful mapper.

Save/replace your existing compare_tables_powerful_auto_mapping.py with this file
(or run this file directly). It should no longer raise the MergedCell AttributeError.
"""

import os
import re
import difflib
import math
import traceback
from typing import List, Tuple, Dict
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from db_config import get_mssql_connection, get_postgres_connection, print_header, get_output_path

# -------------------------
# Matching utilities
# -------------------------
RE_NON_ALNUM_UNDERSCORE = re.compile(r'[^0-9a-zA-Z_]+')
RE_DIGITS = re.compile(r'\d+$')
RE_COL_LETTERS = re.compile(r'^([A-Z]+)')  # for extracting column letters from coordinate

COMMON_PREFIXES = ('is_', 'has_', 'the_', 'tbl_', 'fk_')
COMMON_SUFFIXES = ('_flag', '_yn')

# small mapping of token synonyms that often occur across systems
TOKEN_EQUIV = {
    'desc': 'description',
    'descr': 'description',
    'code': 'code',
    'cd': 'code',
    'id': 'id',
    'pk': 'id',
    'uid': 'id',
    'name': 'name',
    'nm': 'name',
    'created': 'created',
    'createdby': 'created_by',
    'created_on': 'created_date',
    'modified': 'modified',
    'modifiedby': 'modified_by',
    'deleted': 'deleted',
    'company': 'company',
    'client': 'client',
    'sap': 'sap',
    'inco': 'incoterm',
}

def normalize(name: str) -> str:
    """Lowercase, replace non-alnum with underscore, collapse repeated underscores."""
    if not name:
        return ""
    s = RE_NON_ALNUM_UNDERSCORE.sub('_', name.strip()).lower()
    s = re.sub(r'__+', '_', s)
    s = s.strip('_')
    return s

def remove_underscores(s: str) -> str:
    return s.replace('_', '') if s else ''

def strip_common_affixes(s: str) -> str:
    if not s:
        return ''
    for p in COMMON_PREFIXES:
        if s.startswith(p):
            s = s[len(p):]
    for suf in COMMON_SUFFIXES:
        if s.endswith(suf):
            s = s[:-len(suf)]
    return s

def split_tokens(name: str) -> List[str]:
    """Split on underscores, digits boundaries and camelCase; return normalized tokens."""
    if not name:
        return []
    s = normalize(name)
    # split underscores
    parts = [p for p in s.split('_') if p]
    tokens = []
    for part in parts:
        # split alpha/digit boundaries
        subparts = re.findall(r'[A-Za-z]+|\d+', part)
        for sp in subparts:
            tokens.append(TOKEN_EQUIV.get(sp, sp))
    return tokens

def token_overlap_score(a_tokens: List[str], b_tokens: List[str]) -> float:
    if not a_tokens or not b_tokens:
        return 0.0
    set_a = set(a_tokens)
    set_b = set(b_tokens)
    inter = set_a.intersection(set_b)
    avg_len = (len(set_a) + len(set_b)) / 2.0
    if avg_len == 0:
        return 0.0
    return len(inter) / avg_len

def ngram_set(s: str, n: int = 3) -> set:
    s2 = re.sub(r'[^a-z0-9]', '', s.lower())
    if len(s2) < n:
        return {s2} if s2 else set()
    return {s2[i:i+n] for i in range(len(s2)-n+1)}

def jaccard(a: set, b: set) -> float:
    if not a or not b:
        return 0.0
    inter = len(a.intersection(b))
    union = len(a.union(b))
    return inter / union if union > 0 else 0.0

# -------------------------
# Composite scoring
# -------------------------
def score_pair(pg_name: str, m_name: str) -> Tuple[float, Dict]:
    detail = {}
    if not pg_name or not m_name:
        return 0.0, detail

    pg_norm = normalize(pg_name)
    m_norm = normalize(m_name)

    # early exacts
    if pg_norm == m_norm:
        detail['method'] = 'Exact'
        detail['components'] = {'exact': 1.0}
        return 1.0, detail

    if remove_underscores(pg_norm) == remove_underscores(m_norm):
        detail['method'] = 'UnderscoreRemoved'
        detail['components'] = {'underscore_removed': 0.98}
        return 0.98, detail

    # tokens
    pg_tokens = split_tokens(pg_name)
    m_tokens = split_tokens(m_name)
    tok_overlap = token_overlap_score(pg_tokens, m_tokens)
    detail['tok_overlap'] = tok_overlap

    # numeric-suffix awareness
    pg_num = RE_DIGITS.search(pg_norm)
    m_num = RE_DIGITS.search(m_norm)
    num_bonus = 0.0
    if pg_num and m_num and pg_num.group() == m_num.group():
        num_bonus = 0.05

    # substring / prefix / suffix
    substr_flag = 0
    if pg_norm in m_norm or m_norm in pg_norm:
        substr_flag = 0.75

    # n-gram jaccard (3-gram)
    pg_ngrams = ngram_set(pg_norm, 3)
    m_ngrams = ngram_set(m_norm, 3)
    ng_jacc = jaccard(pg_ngrams, m_ngrams)
    detail['ngram_jaccard'] = ng_jacc

    # difflib ratio (sequence)
    seq_ratio = difflib.SequenceMatcher(None, pg_norm, m_norm).ratio()
    detail['seq_ratio'] = seq_ratio

    # token partial / abbreviation handling:
    token_subscore = 0.0
    for at in pg_tokens:
        for bt in m_tokens:
            if at == bt:
                token_subscore = max(token_subscore, 1.0)
            elif at in bt or bt in at:
                token_subscore = max(token_subscore, 0.8)
            else:
                token_subscore = max(token_subscore, difflib.SequenceMatcher(None, at, bt).ratio() * 0.6)

    detail['token_subscore'] = token_subscore

    # Weighted combination
    w_token_overlap = 0.35
    w_token_sub = 0.20
    w_seq = 0.18
    w_ngram = 0.12
    w_substr = 0.10
    base_score = (
        w_token_overlap * tok_overlap
        + w_token_sub * token_subscore
        + w_seq * seq_ratio
        + w_ngram * ng_jacc
        + w_substr * (1.0 if substr_flag else 0.0)
    )

    base_score = min(1.0, base_score + num_bonus)
    score = round(base_score, 4)

    if tok_overlap >= 0.8 or token_subscore >= 0.95:
        method = f"TokenStrong({tok_overlap:.2f})"
    elif tok_overlap >= 0.4:
        method = f"Token({tok_overlap:.2f})"
    elif ng_jacc >= 0.45:
        method = f"NGram({ng_jacc:.2f})"
    elif seq_ratio >= 0.75:
        method = f"Fuzzy({seq_ratio:.2f})"
    elif substr_flag:
        method = "Substring"
    else:
        method = f"FuzzyLow({seq_ratio:.2f})"

    detail['method'] = method
    detail['components'] = {
        'token_overlap': tok_overlap,
        'token_subscore': token_subscore,
        'seq_ratio': seq_ratio,
        'ngram_jacc': ng_jacc,
        'substr_flag': substr_flag,
        'num_bonus': num_bonus,
    }
    return score, detail

def interpret_note(score: float) -> str:
    if score >= 0.95:
        return "High-confidence match"
    if score >= 0.75:
        return "Probable match — please review"
    if score >= 0.35:
        return "Low confidence — manual review recommended"
    return ""

# -------------------------
# Mapping orchestration
# -------------------------
def suggest_mappings(pg_df: pd.DataFrame, mssql_df: pd.DataFrame, threshold: float = 0.35, one_to_one: bool = True):
    pg_cols = [str(x) for x in pg_df['column_name'].tolist()]
    m_cols = [str(x) for x in mssql_df['COLUMN_NAME'].tolist()]

    # Precompute all pair scores
    all_scores = {}
    for p in pg_cols:
        all_scores[p] = {}
        for m in m_cols:
            sc, detail = score_pair(p, m)
            all_scores[p][m] = (sc, detail)

    chosen_m_for_p = {p: ('', 0.0, {}) for p in pg_cols}

    if one_to_one:
        flat = []
        for p in pg_cols:
            for m in m_cols:
                sc, det = all_scores[p][m]
                if sc > 0:
                    flat.append((p, m, sc, det))
        flat.sort(key=lambda x: x[2], reverse=True)
        used_m = set()
        for p, m, sc, det in flat:
            if chosen_m_for_p[p][1] >= sc:
                continue
            if m in used_m:
                continue
            chosen_m_for_p[p] = (m, sc, det)
            used_m.add(m)
    else:
        for p in pg_cols:
            best_m = ''
            best_sc = 0.0
            best_det = {}
            for m in m_cols:
                sc, det = all_scores[p][m]
                if sc > best_sc:
                    best_sc = sc
                    best_m = m
                    best_det = det
            chosen_m_for_p[p] = (best_m, best_sc, best_det)

    mapping_rows = []
    diagnostics = {}
    for p in pg_cols:
        m, sc, det = chosen_m_for_p[p]
        diagnostics[p] = {'mssql_candidate': m, 'score': sc, 'detail': det}
        if sc >= threshold:
            mapping_rows.append({
                'PG_COLUMN_NAME': p,
                'MSSQL_COLUMN_NAME': m,
                'SUGGESTED_SCORE': round(sc, 2),
                'NOTES': interpret_note(sc),
                '_MATCH_METHOD': det.get('method', '') if det else ''
            })
        else:
            mapping_rows.append({
                'PG_COLUMN_NAME': p,
                'MSSQL_COLUMN_NAME': '-',
                'SUGGESTED_SCORE': 0,
                'NOTES': 'not find',
                '_MATCH_METHOD': det.get('method', '') if det else ''
            })
    return mapping_rows, diagnostics

# -------------------------
# Excel output (fixed for MergedCell)
# -------------------------
from openpyxl import load_workbook

def _col_letters_from_cell(cell) -> str:
    """
    Safely extract column letters from any cell-like object using its coordinate.
    Works for normal Cell and MergedCell.
    """
    coord = getattr(cell, "coordinate", None)
    if not coord:
        return ""
    m = RE_COL_LETTERS.match(coord)
    return m.group(1) if m else ""

def write_excel(output_file: str, comp_df: pd.DataFrame, mapping_rows: List[dict]):
    os.makedirs(os.path.dirname(output_file) or '.', exist_ok=True)
    df_map = pd.DataFrame(mapping_rows, columns=['PG_COLUMN_NAME', 'MSSQL_COLUMN_NAME', 'SUGGESTED_SCORE', 'NOTES', '_MATCH_METHOD'])
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        comp_df.to_excel(writer, sheet_name='Comparison', index=False, startrow=1)
        df_map.to_excel(writer, sheet_name='AutoMapping', index=False, startrow=1, columns=['PG_COLUMN_NAME', 'MSSQL_COLUMN_NAME', 'SUGGESTED_SCORE', 'NOTES'])
        workbook = writer.book
        ws_comp = writer.sheets['Comparison']
        ws_map = writer.sheets['AutoMapping']

        # styling
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        col_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        col_font = Font(bold=True, size=11)
        center = Alignment(horizontal='center', vertical='center')
        thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Comparison header
        ws_comp['A1'] = 'mssql'
        ws_comp['D1'] = 'postgres'
        ws_comp.merge_cells('A1:B1')
        ws_comp.merge_cells('D1:E1')
        for c in ['A1','D1']:
            ws_comp[c].fill = header_fill
            ws_comp[c].font = header_font
            ws_comp[c].alignment = center
        # headers row 2
        ws_comp['A2'] = 'COLUMN_NAME'
        ws_comp['B2'] = 'DATA_TYPE'
        ws_comp['C2'] = ''
        ws_comp['D2'] = 'column_name'
        ws_comp['E2'] = 'data_type'
        for c in ['A2','B2','D2','E2']:
            ws_comp[c].fill = col_fill
            ws_comp[c].font = col_font
            ws_comp[c].alignment = center
        ws_comp.column_dimensions['A'].width = 30
        ws_comp.column_dimensions['B'].width = 20
        ws_comp.column_dimensions['C'].width = 3
        ws_comp.column_dimensions['D'].width = 30
        ws_comp.column_dimensions['E'].width = 20

        max_rows = comp_df.shape[0]
        # Use safe column extraction to avoid MergedCell.column_letter error
        for row in ws_comp.iter_rows(min_row=1, max_row=max_rows + 2, min_col=1, max_col=5):
            for cell in row:
                col_letters = _col_letters_from_cell(cell)
                if col_letters != 'C':
                    cell.border = thin
                    cell.alignment = Alignment(vertical='center')
        # separator gray
        gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        for r in range(1, max_rows + 3):
            ws_comp[f'C{r}'].fill = gray_fill

        # AutoMapping header
        ws_map['A1'] = 'Auto mapping (PG -> MSSQL)'
        ws_map.merge_cells('A1:D1')
        ws_map['A1'].fill = header_fill
        ws_map['A1'].font = header_font
        ws_map['A1'].alignment = center

        # column headers row2
        headers = ['PG_COLUMN_NAME', 'MSSQL_COLUMN_NAME', 'SUGGESTED_SCORE', 'NOTES']
        for i, h in enumerate(headers, start=1):
            cell = ws_map.cell(row=2, column=i, value=h)
            cell.fill = col_fill
            cell.font = col_font
            cell.alignment = center

        ws_map.column_dimensions['A'].width = 40
        ws_map.column_dimensions['B'].width = 40
        ws_map.column_dimensions['C'].width = 12
        ws_map.column_dimensions['D'].width = 60

        map_rows = df_map.shape[0]
        # Safe border/alignment loop for AutoMapping too
        for row in ws_map.iter_rows(min_row=1, max_row=map_rows + 2, min_col=1, max_col=4):
            for cell in row:
                # extract col letters safely
                col_letters = _col_letters_from_cell(cell)
                cell.border = thin
                if col_letters == 'D':
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                else:
                    cell.alignment = Alignment(vertical='center')

        # Ensure SUGGESTED_SCORE are numeric where present and formatted to 2 decimals
        for row_idx in range(3, 3 + map_rows):
            cell = ws_map.cell(row=row_idx, column=3)  # column C
            if cell.value is not None and cell.value != '':
                try:
                    num = float(cell.value)
                    cell.value = num
                    cell.number_format = '0.00'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                except ValueError:
                    # leave as-is
                    pass

    print("Saved Excel:", os.path.exists(output_file), " Size:", os.path.getsize(output_file) if os.path.exists(output_file) else None)

# -------------------------
# Main CLI
# -------------------------
def main():
    print_header("Powerful PG->MSSQL Auto-Mapping (fixed merged-cell handling)")
    mssql_conn = None
    pg_conn = None
    try:
        print("\n[1/3] Connecting to MSSQL...")
        mssql_conn = get_mssql_connection()
        print("✓ MSSQL Connected")
        print("\n[2/3] Connecting to PostgreSQL...")
        pg_conn = get_postgres_connection()
        print("✓ PostgreSQL Connected")

        print("\n" + "="*80)
        m_input = input("Enter MSSQL table name (or schema.table) : ").strip()
        pg_table = input("Enter PostgreSQL table name (table only or schema.table): ").strip()
        print("="*80)

        # parse mssql schema
        if '.' in m_input:
            m_schema, m_table = [p.strip() for p in m_input.split('.', 1)]
        else:
            m_schema = 'dbo'
            m_table = m_input

        # parse pg schema optionally
        if '.' in pg_table:
            p_schema, p_table = [p.strip() for p in pg_table.split('.', 1)]
        else:
            p_schema = 'public'
            p_table = pg_table

        print("\n[3/3] Fetching table structures...")
        mssql_q = f"""
            SELECT COLUMN_NAME, DATA_TYPE
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_NAME = '{m_table}' AND TABLE_SCHEMA = '{m_schema}'
            ORDER BY ORDINAL_POSITION
        """
        pg_q = f"""
            SELECT column_name, data_type
            FROM information_schema.columns
            WHERE table_name = '{p_table}' AND table_schema = '{p_schema}'
            ORDER BY ordinal_position
        """
        df_mssql = pd.read_sql(mssql_q, mssql_conn)
        if df_mssql.empty:
            print(f"✗ MSSQL table '{m_table}' in schema '{m_schema}' not found or no columns visible.")
            return
        df_pg = pd.read_sql(pg_q, pg_conn)
        if df_pg.empty:
            print(f"✗ PostgreSQL table '{p_table}' in schema '{p_schema}' not found or no columns visible.")
            return

        print(f"✓ MSSQL columns: {len(df_mssql)}")
        print(f"✓ PostgreSQL columns: {len(df_pg)}")

        # Build comparison DataFrame (side-by-side)
        max_rows = max(len(df_mssql), len(df_pg))
        rows = []
        for i in range(max_rows):
            r = {}
            if i < len(df_mssql):
                r['MSSQL_COLUMN_NAME'] = df_mssql.iloc[i]['COLUMN_NAME']
                r['MSSQL_DATA_TYPE'] = df_mssql.iloc[i]['DATA_TYPE']
            else:
                r['MSSQL_COLUMN_NAME'] = ''
                r['MSSQL_DATA_TYPE'] = ''
            r['SEPARATOR'] = ''
            if i < len(df_pg):
                r['PG_COLUMN_NAME'] = df_pg.iloc[i]['column_name']
                r['PG_DATA_TYPE'] = df_pg.iloc[i]['data_type']
            else:
                r['PG_COLUMN_NAME'] = ''
                r['PG_DATA_TYPE'] = ''
            rows.append(r)
        df_comp = pd.DataFrame(rows)

        # Suggest mappings
        threshold = 0.35
        mapping_rows, diagnostics = suggest_mappings(df_pg, df_mssql, threshold=threshold, one_to_one=True)

        # Save to Excel
        output_file = get_output_path(f"Compare_{m_table}_vs_{p_table}_powerful_mapping_fixed.xlsx")
        print("Saving to:", output_file)
        write_excel(output_file, df_comp, mapping_rows)

        print("\nMapping complete. Open the AutoMapping sheet to review suggestions.")
        for i, p in enumerate(df_pg['column_name'].tolist()[:10]):
            diag = diagnostics.get(p, {})
            print(f"{p} => {diag.get('mssql_candidate','')} (score={diag.get('score',0):.2f}) method={diag.get('detail',{}).get('method','')}")
    except Exception as e:
        print("✗ Failed:", e)
        traceback.print_exc()
    finally:
        try:
            if mssql_conn:
                mssql_conn.close()
        except Exception:
            pass
        try:
            if pg_conn:
                pg_conn.close()
        except Exception:
            pass
        input("\nPress Enter to exit...")

if __name__ == '__main__':
    main()